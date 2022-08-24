from __future__ import print_function

import os
from tkinter import *
from tkinter.ttk import Progressbar
from tkinter import ttk, messagebox
from tkinter import filedialog as fd

import json
from datetime import datetime
import calendar
from openpyxl import Workbook, load_workbook
import pandas as pd

from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials


def main_win():
    # Create an instance of tkinter window
    win = Tk()
    win.title('NOC Ninja')
    win.iconbitmap('hard-work.ico')

    # Set the size of the tkinter window
    width = 900
    height = 450
    # set the screen location
    screen_w = win.winfo_screenwidth()
    screen_h = win.winfo_screenheight()
    x = (screen_w / 2) - (width / 2)
    y = (screen_h / 2) - (height / 2)
    win.geometry("%dx%d+%d+%d" % (width, height, x, y))
    # creating a tree frame
    tree_frame = Frame(win)
    tree_frame.pack(pady=20, padx=10)
    # creating a button frame
    btn_frame = Frame(win)
    btn_frame.pack(pady=10, padx=10)

    # Create an object of Style widget
    style = ttk.Style()
    style.theme_use('clam')

    # Constructing vertical scrollbar with treeview
    tree_scroll = Scrollbar(tree_frame)
    tree_scroll.pack(side=RIGHT, fill=Y)

    # Add a Treeview widget
    tree = ttk.Treeview(tree_frame, column=("Date", "Day", "Hours", "Name"), show='headings',
                        yscrollcommand=tree_scroll.set, height=15)
    tree.pack()
    tree_scroll.config(command=tree.yview)

    tree.column("# 1", anchor=CENTER)
    tree.heading("# 1", text="Date")
    tree.column("# 2", anchor=CENTER)
    tree.heading("# 2", text="Day")
    tree.column("# 3", anchor=CENTER)
    tree.heading("# 3", text="Hours")
    tree.column("# 4", anchor=CENTER)
    tree.heading("# 4", text="Name")

    ########################################################################################
    # defining functions
    ########################################################################################

    def select_file():
        global filenm
        global filtered
        filtered = False

        filetypes = (
            ('Excel files', '*.xlsx'),
            ('All files', '*.*')
        )

        filename = fd.askopenfilename(
            title='Open a file',
            initialdir='/',
            filetypes=filetypes)
        filenm = str(filename)

        if filenm == '':
            messagebox.showinfo("showinfo", "No file was picked!")
            return
        # trimming the file path to get the file name
        fn_index = len(filenm.split("/")) - 1
        fn = filenm.split("/")[fn_index]
        file_Label.configure(text="File loaded: " + fn)

        # load a work book and worksheet
        wb = load_workbook(filenm,data_only=True)
        ws = wb.active

        # creating shift dictionaries by extracting data from col A, B, C and I on
        # the excel sheet  and appending it to a shifts list
        shiftslist = []
        for row in range(2, 95):
            # formatting the dates
            date_cell = str(ws['B' + str(row)].value)

            if date_cell != 'None':
                DT = datetime.strptime(date_cell, '%Y-%m-%d %H:%M:%S')
                formatted_date = DT.strftime('%Y-%m-%d')

            # Stop Condition whenever hits 'none' on 'employee' --> end of the month
            if str(ws['H' + str(row)].value) == 'None':
                break
            # creating shifts list
            shift = {
                "date": formatted_date,
                "day": calendar.day_name[DT.weekday()],
                "hours": ws['C' + str(row)].value,
                "employee": ws['H' + str(row)].value,
                "prev_shift": ws['H' + str(row - 1)].value,
                "next_shift": ws['H' + str(row + 1)].value

            }
            shiftslist.append(shift)

        # creating a json shifts obj and Writing it to a file
        shifts_json = json.dumps(shiftslist)

        shifts_file = open("all_shifts.json", "w")
        shifts_file.write(shifts_json)
        shifts_file.close()
        # Opening JSON file
        f = open('all_shifts.json')

        # returns JSON object as a dictionary
        global data
        data = json.load(f)

        pressed_ok = messagebox.askokcancel("Load", f"Continue loading {fn}?")
        if pressed_ok: get_shifts()

    ########################################################################################
    # Load btn
    ########################################################################################
    def get_shifts():
        # boolean that indicates if all shifts were loaded or filtered shifts were loaded
        global filtered
        filtered = False

        try:
            data
        except NameError:
            messagebox.showerror("Error", "No file was loaded!")
            return
        # delete previous data from the table
        for i in tree.get_children():
            tree.delete(i)
        win.update()
        # Iterating through the json list and  Insert the data in Treeview widget
        for i in data:
            tree.insert('', 'end', text="1", values=(i['date'], i['day'], i['hours'], i['employee']))

    ########################################################################################
    # Filter btn
    ########################################################################################
    def get_employee_shifts():
        # boolean that indicates if all shifts were loaded or filtered shifts were loaded
        global filtered
        filtered = True

        try:
            data
        except NameError:
            messagebox.showerror("Error", "No data to filter!")
            return
        # delete previous data from the table
        for i in tree.get_children():
            tree.delete(i)
        win.update()

        global filtered_shifts
        filtered_shifts = []
        # iterating the data to filter employee match
        for i in data:
            if i['employee'].lower() == EMPLOYEE_NAME.get().strip().lower():
                tree.insert('', 'end', text="1", values=(i['date'], i['day'], i['hours'], i['employee']))
                # creating filtered shift list by employee
                shift = {
                    "date": i['date'],
                    "day": i['day'],
                    "hours": i['hours'],
                    "employee": i['employee'],
                    "prev_shift": i['prev_shift'],
                    "next_shift": i['next_shift']
                }
                filtered_shifts.append(shift)
        if not filtered_shifts: tree.insert('', 'end', text="1",
                                            values=("no match", "no match", "no match", "no match"))
        print(filtered_shifts)

    ########################################################################################
    # Save btn
    ########################################################################################
    def save():
        try:
            data
        except NameError:
            messagebox.showerror("Error", "No data loaded!")
            return
        # creating pandas data frame from all shifts
        if not filtered:
            df = pd.read_json(json.dumps(data))

        else:
            try:
                if not filtered_shifts:
                    messagebox.showerror("Error", "no data was filter to save!")
                    return
            except NameError:
                messagebox.showerror("Error", "No data to save!")
                return
            # creating pandas data frame from all shifts
            df = pd.read_json(json.dumps(filtered_shifts))

        file = fd.asksaveasfile(defaultextension='.txt',
                                filetypes=[
                                    ("Text file", '.txt'),
                                    ("HTML file", '.html'),
                                ])
        # to implement - calculate total hours and estimated salary
        sum = len(filtered_shifts) * 8
        sal = sum * 38

        if file is not None:
            file.write(str(df) + '\n\n')
            file.write(f"Total hours:{sum} \n")
            file.write(f"estimated salary:{sal} \n")
            file.close()
            messagebox.showinfo("showinfo", "File saved!")

    ########################################################################################
    # request the  Calendar API
    ########################################################################################

    email_dictionary = {
        'tamir padlad': "tamir.padlad@gmail.com",
        'tal tesler': "tal@tesler",
        'gilad aslan': "gilad@aslan",
        'michael afonin': "Shaharcch@cc",
        'eitan goldstein': "eitan@goldstein",
        'shay maatuk': "Shay.maatuk@gmail.com"
    }


    def write_to_calendar():

        try:
            # checking the bool flag to determine what data is loaded to the window
            if not filtered:
                print("uploading all shifts")

        except NameError:
            messagebox.showerror("Error", "No data to upload!")
            return



        # checking the bool flag to determine what data is loaded to the window
        if not filtered:  # all shifts
            shifts = data
            print(shifts)
            pressed_ok = messagebox.askokcancel("Upload", "Send calendar invitaions to all employees?")
            if not pressed_ok: return
        else:  # filter by employee
            messagebox.showerror("Error", "cannot upload to filtered shifts!")
            return



        SCOPES = ['https://www.googleapis.com/auth/calendar']
        creds = None
        # The file token.json stores the user's access and refresh tokens, and is
        # created automatically when the authorization flow completes for the first
        # time.
        if os.path.exists('token.json'):
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        # If there are no (valid) credentials available, let the user log in.
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
            ##################################################################################################################################################################################
                # It will start a local web server to listen for the authorization response.
                # Once authorization is complete the authorization server will redirect the user’s browser
                # to the local web server. The web server will get the authorization code from the response and shutdown
                try:
                    flow = InstalledAppFlow.from_client_secrets_file('Desktop NOC_ninja oauth client 1.json', SCOPES)
                    creds = flow.run_local_server(port=0)
                except:
                    Exception
                print(Exception)
            ##################################################################################################################################################################################
            # Save the credentials for the next run
            with open('token.json', 'w') as token:
                token.write(creds.to_json())

        try:
            # Call the Calendar API
            service = build('calendar', 'v3', credentials=creds)

            # Iterating the filtered shift to pull the start and end time of the shift
            for shift in shifts:
                start_time = f"{shift['date']}T{shift['hours'].split('-')[0]}:00+03:00"

                # checking if a shifts starts on one day and ends the day after
                if shift['hours'].split('-')[1].replace(" ", "") == '01:00':
                    #if its last day of the month set end date to 1st of the nexxt month
                    month_last_day = calendar.monthrange(int(shift['date'].split("-")[0]), int(shift['date'].split("-")[1]))[1]
                    if month_last_day == int(shift['date'].split("-")[2]):
                        new_date = str(shift['date'].split("-")[0]) + "-" + str(int(shift['date'].split("-")[1] )+ 1) + "-" + '1'

                    #if not the last  day of the month set end date to day after
                    else:
                        new_date = str(shift['date'].split("-")[0]) + "-" + str(shift['date'].split("-")[1]) + "-" + str(int(shift['date'].split("-")[2]) + 1)
                    end_time = f"{new_date}T{shift['hours'].split('-')[1]}:00+03:00"
                else:
                    end_time = f"{shift['date']}T{shift['hours'].split('-')[1]}:00+03:00"

                #inviting the relevant user
                employee_mail = email_dictionary.get(shift.get('employee').lower())

                event = {
                    'summary': f"{shift.get('employee')} - PCOC On-call",

                    'start': {
                        'dateTime': start_time.replace(" ", ""),
                    },
                    'end': {
                        'dateTime': end_time.replace(" ", ""),
                    },
                    'attendees': [
                        {'email': employee_mail},

                    ],

                    'reminders': {
                        'useDefault': False,
                        'overrides': [
                            {'method': 'email', 'minutes': 24 * 60},
                            {'method': 'popup', 'minutes': 15},
                        ],
                    },
                    'description': f"contact {shift['prev_shift']} for shift handover, when finishing handover to {shift['next_shift']} "

                }

                event = service.events().insert(calendarId='primary', body=event).execute()
                print('Event created: %s' % (event.get('htmlLink')))

            messagebox.showinfo("showinfo", "New shifts were uploaded to NOC calender!")


        except HttpError as error:
             print('An error occurred: %s' % error)

    ########################################################################################
    # Buttons
    ########################################################################################
    open_file_btn = ttk.Button(btn_frame, text="Open file", command=select_file)
    open_file_btn.grid(row=1, column=0, padx=5)

    load_btn = ttk.Button(btn_frame, text="All shifts", command=get_shifts)
    load_btn.grid(row=1, column=1, padx=5)

    filter_btn = ttk.Button(btn_frame, text="Filter", command=get_employee_shifts)
    filter_btn.grid(row=1, column=2, padx=5)

    EMPLOYEE_NAME = StringVar()
    name_entry = ttk.Entry(btn_frame, width=10, font=('Arial', 16), textvariable=EMPLOYEE_NAME)
    name_entry.bind("<Return>", lambda e: get_employee_shifts())
    name_entry.grid(row=1, column=3, padx=5)

    save_btn = ttk.Button(btn_frame, text="Save", command=save)
    save_btn.grid(row=1, column=4, padx=5)

    up_btn = ttk.Button(btn_frame, text="Upload to calender", command=write_to_calendar)
    up_btn.grid(row=1, column=5, padx=5)

    file_Label = Label(win, text="Select a file to get started!", fg='black')
    file_Label.config(font=('Calibri (Body)', 10))
    file_Label.place(x=20, y=420)

    win.mainloop()


#######################################################################################################################
# splash screen
#######################################################################################################################
# Create and set the size of the tkinter window
splash_win = Tk()
width_of_window = 900
height_of_window = 450
screen_width = splash_win.winfo_screenwidth()
screen_height = splash_win.winfo_screenheight()
x_coordinate = (screen_width / 2) - (width_of_window / 2)
y_coordinate = (screen_height / 2) - (height_of_window / 2)
splash_win.geometry("%dx%d+%d+%d" % (width_of_window, height_of_window, x_coordinate, y_coordinate))
splash_win.overrideredirect(1)

s = ttk.Style()
s.theme_use('clam')
s.configure("red.Horizontal.TProgressbar", foreground='red', background='#4f4f4f')
progress = Progressbar(splash_win, style="red.Horizontal.TProgressbar", orient=HORIZONTAL, length=1000,
                       mode='determinate', )


def bar():
    l4 = Label(splash_win, text='Loading...', fg='black', bg='#e1e9e1')
    lst4 = ('Calibri (Body)', 10)
    l4.config(font=lst4)
    l4.place(x=20, y=400)

    import time
    r = 0
    for i in range(100):
        progress['value'] = r
        splash_win.update_idletasks()
        time.sleep(0.02)
        r = r + 1

    splash_win.destroy()
    main_win()


progress.place(x=-10, y=435)

Frame(splash_win, width=900, height=440, bg='#e1e9e1').place(x=0, y=0)

start_btn = Button(splash_win, width=10, height=1, text='Get Started', command=bar, border=0.5, fg='black',
                   bg='#e1e9e1')
start_btn.place(x=750, y=400)

# Add image file
bg = PhotoImage(file="hard-work-main.png")
# Show image using label
label1 = Label(splash_win, image=bg, bg='#e1e9e1')
label1.place(x=310, y=70)
# label
l1 = Label(splash_win, text="NOC Ninja", fg='black', bg='#e1e9e1')
lst1 = ('Calibri (Body)', 25, 'bold')
l1.config(font=lst1)
l1.place(x=350, y=335)

splash_win.mainloop()
